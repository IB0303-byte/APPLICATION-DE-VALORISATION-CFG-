import streamlit as st
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import io
import datetime
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import base64
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="CFG Bank – Valorisation OPCVM",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #f4f6fa; }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
    }
    [data-testid="stSidebar"] * { color: #e0e0e0 !important; }
    [data-testid="stSidebar"] hr { border-color: #e31e24 !important; }
    .cfg-header {
        background: linear-gradient(90deg, #1a1a2e 0%, #e31e24 100%);
        padding: 18px 30px; border-radius: 12px;
        display: flex; align-items: center; gap: 24px;
        margin-bottom: 24px; box-shadow: 0 4px 20px rgba(227,30,36,.25);
    }
    .cfg-header h1 { color:#fff; margin:0; font-size:1.55rem; font-weight:700; }
    .cfg-header p  { color:#ffd0d0; margin:4px 0 0; font-size:.85rem; }
    .kpi-row { display:flex; gap:16px; margin-bottom:20px; flex-wrap:wrap; }
    .kpi-card {
        background:#fff; border-radius:12px; padding:18px 22px;
        flex:1; min-width:160px; box-shadow:0 2px 12px rgba(0,0,0,.07);
        border-top:4px solid #e31e24;
    }
    .kpi-card .kpi-label { font-size:.78rem; color:#888; text-transform:uppercase; letter-spacing:.5px; }
    .kpi-card .kpi-value { font-size:1.5rem; font-weight:700; color:#1a1a2e; margin-top:4px; }
    .kpi-card .kpi-sub   { font-size:.78rem; color:#aaa; margin-top:2px; }
    .section-title {
        font-size:1.05rem; font-weight:700; color:#1a1a2e;
        border-left:4px solid #e31e24; padding-left:10px; margin:20px 0 12px;
    }
    .stButton > button {
        background: linear-gradient(90deg,#e31e24,#c0141a);
        color:#fff !important; border:none; border-radius:8px;
        font-weight:600; padding:10px 22px;
        box-shadow:0 3px 10px rgba(227,30,36,.3);
    }
    .stButton > button:hover { opacity:.88; }
    [data-testid="stTabs"] [role="tab"] {
        font-weight:600; font-size:.9rem; color:#555;
        border-bottom:3px solid transparent; padding:10px 20px;
    }
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        color:#e31e24 !important; border-bottom:3px solid #e31e24 !important;
    }
    .stDownloadButton > button {
        background:linear-gradient(90deg,#0f3460,#16213e);
        color:#fff !important; border:none; border-radius:8px;
        font-weight:600; padding:10px 22px;
    }
    [data-testid="stMetric"] {
        background:#fff; border-radius:10px; padding:12px;
        box-shadow:0 2px 8px rgba(0,0,0,.06);
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# LOGO
# ─────────────────────────────────────────────
def get_logo_b64(path="LOGO1.png"):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return None

logo_b64 = get_logo_b64()
logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:52px;border-radius:6px;"/>' if logo_b64 else "🏦"

st.markdown(f"""
<div class="cfg-header">
    {logo_html}
    <div>
        <h1>Système de Valorisation OPCVM</h1>
        <p>Conforme à la Circulaire CDVM N°02/04 – Titres de créances à taux fixe remboursables in fine</p>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div style="background:#fff; border-left:4px solid #e31e24; border-radius:8px;
            padding:10px 20px; margin-bottom:18px; box-shadow:0 1px 6px rgba(0,0,0,.07);
            display:flex; align-items:center; gap:32px; flex-wrap:wrap;">
    <span style="font-size:.82rem; color:#555;">
        <b style="color:#1a1a2e;">ℹ️ En cas de besoin, contacter :</b>
        &nbsp;&nbsp;
        <b>JEAN MATA</b>
        &nbsp;|&nbsp;
        📞 <a href="tel:+33987238238" style="color:#e31e24;text-decoration:none;">+33 987 238 238</a>
        &nbsp;|&nbsp;
        ✉️ <a href="mailto:jean@gmail.com" style="color:#e31e24;text-decoration:none;">jean@gmail.com</a>
    </span>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
for key in ["bam_curve", "df_titres", "df_results"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    if logo_b64:
        st.markdown(
            f'<div style="text-align:center;padding:10px 0 20px">'
            f'<img src="data:image/png;base64,{logo_b64}" style="height:44px;"/></div>',
            unsafe_allow_html=True
        )
    st.markdown("---")
    st.markdown("### 🗂️ Navigation")
    page = st.radio("", [
        "🏠 Accueil & Courbe BAM",
        "📊 Valorisation",
        "📈 Analyse & Graphiques",
        "📋 Résultats & Export"
    ], label_visibility="collapsed")
    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:.82rem;color:#ccc;line-height:1.8">
    <b>Date :</b> {datetime.date.today().strftime('%d/%m/%Y')}<br/>
    <b>Référentiel :</b> Circulaire CDVM N°02/04<br/>
    <b>Source taux :</b> Bank Al-Maghrib
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(
        '<div style="font-size:.75rem;color:#888;text-align:center">'
        "CFG Bank – Direction Gestion d'Actifs<br/>© 2026</div>",
        unsafe_allow_html=True
    )

# ─────────────────────────────────────────────
# CONSTANTES GRAPHIQUES
# ─────────────────────────────────────────────
CFG_RED  = "#e31e24"
CFG_DARK = "#1a1a2e"
CFG_BLUE = "#0f3460"
CFG_MID  = "#16213e"
COLORS   = [CFG_RED, CFG_DARK, CFG_BLUE, CFG_MID, "#888888", "#aaaaaa"]

def chart_layout(fig, title="", h=400, legend=True):
    fig.update_layout(
        title=dict(text=title, font=dict(size=14, color=CFG_DARK), x=0.01),
        height=h,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Arial", size=12, color=CFG_DARK),
        margin=dict(l=60, r=30, t=55, b=60),
        showlegend=legend,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(size=11)),
        xaxis=dict(showgrid=True, gridcolor="#eeeeee", zeroline=False,
                   linecolor="#cccccc", linewidth=1),
        yaxis=dict(showgrid=True, gridcolor="#eeeeee", zeroline=False,
                   linecolor="#cccccc", linewidth=1),
    )
    return fig

# ─────────────────────────────────────────────
# HELPER: INTERPOLATION (Art. 6 Circulaire)
# ─────────────────────────────────────────────
def interpolate_rate(maturity_days, curve_maturities_days, curve_rates):
    mats = np.array(curve_maturities_days, dtype=float)
    rats = np.array(curve_rates, dtype=float)
    idx  = np.argsort(mats)
    mats, rats = mats[idx], rats[idx]
    if maturity_days <= mats[0]:
        return float(rats[0])
    if maturity_days >= mats[-1]:
        return float(rats[-1])
    for i in range(len(mats) - 1):
        if mats[i] <= maturity_days <= mats[i + 1]:
            x0, x1 = mats[i], mats[i + 1]
            y0, y1 = rats[i], rats[i + 1]
            return float(y0 + (y1 - y0) * (maturity_days - x0) / (x1 - x0))
    return float(rats[-1])

def _days_to_years_label(d: datetime.date, eval_date: datetime.date) -> str:
    """Retourne 'DD/MM/YYYY (X.XX ans)' pour affichage des échéances."""
    days  = (d - eval_date).days
    years = days / 365.25
    return f"{d.strftime('%d/%m/%Y')} ({years:.2f} ans)"

# ─────────────────────────────────────────────
# HELPER: PARSE BAM CSV OFFICIEL
# Format: "Date d'échéance";Transaction;"Taux moyen pondéré";"Date de la valeur"
# ─────────────────────────────────────────────
def parse_bam_csv(file_content: bytes, eval_date: datetime.date):
    try:
        text = file_content.decode("utf-8-sig")
    except Exception:
        text = file_content.decode("latin-1")

    lines = text.splitlines()
    mats, rats, labels = [], [], []
    date_valeur = None

    for line in lines:
        line_clean = line.strip().strip('"')
        # Skip header/metadata lines
        if not line_clean:
            continue
        lower = line_clean.lower()
        if any(kw in lower for kw in ["taux de r", "en million", "total", "date d'", "transaction"]):
            continue

        parts = line_clean.split(";")
        if len(parts) < 3:
            continue

        date_str = parts[0].strip().strip('"')
        taux_str = parts[2].strip().strip('"').replace("%", "").replace(",", ".").replace(" ", "")

        # Grab date_valeur from column 4 if present
        if len(parts) >= 4 and date_valeur is None:
            dv = parts[3].strip().strip('"')
            try:
                date_valeur = datetime.datetime.strptime(dv, "%d/%m/%Y").date()
            except Exception:
                pass

        try:
            d    = datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            rate = float(taux_str) / 100
            days = (d - eval_date).days
            if days <= 0:
                continue
            label = _days_to_years_label(d, eval_date)
            mats.append(days)
            rats.append(rate)
            labels.append(label)
        except Exception:
            continue

    if len(rats) < 2:
        raise ValueError(
            "Pas assez de points parsés depuis le CSV BAM. "
            "Vérifiez que la date d'évaluation est antérieure aux échéances du fichier."
        )
    return mats, rats, labels, date_valeur or eval_date

# ─────────────────────────────────────────────
# HELPER: FETCH BAM via ANTHROPIC API (proxy)
# L'API Anthropic est accessible depuis Streamlit Cloud.
# On demande à Claude de récupérer et retourner la courbe BAM.
# ─────────────────────────────────────────────

_BAM_PAGE_URL = (
    "https://www.bkam.ma/Marches/Principaux-indicateurs/Marche-obligataire/"
    "Marche-des-bons-de-tresor/Marche-secondaire/Taux-de-reference-des-bons-du-tresor"
)

def _parse_bam_html_table(html_text: str, eval_date: datetime.date):
    """Parse une page HTML BAM et extrait la courbe de taux."""
    soup   = BeautifulSoup(html_text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        raise ValueError("Aucune table trouvée dans la réponse HTML.")
    mats, rats, labels = [], [], []
    date_valeur = None
    for table in tables:
        for row in table.find_all("tr"):
            cols = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if len(cols) < 3:
                continue
            try:
                d      = datetime.datetime.strptime(cols[0].strip(), "%d/%m/%Y").date()
                taux_r = cols[2].strip().replace("%","").replace(",",".").replace(" ","")
                rate   = float(taux_r) / 100
                days   = (d - eval_date).days
                if days <= 0:
                    continue
                if len(cols) >= 4 and date_valeur is None:
                    try:
                        date_valeur = datetime.datetime.strptime(cols[3].strip(), "%d/%m/%Y").date()
                    except Exception:
                        pass
                labels.append(_days_to_years_label(d, eval_date))
                mats.append(days)
                rats.append(rate)
            except Exception:
                continue
        if len(rats) >= 2:
            break
    if len(rats) < 2:
        raise ValueError("Impossible de parser les taux depuis le HTML.")
    return mats, rats, labels, date_valeur or eval_date


def fetch_bam_via_anthropic(eval_date: datetime.date, api_key: str):
    """
    Utilise l'API Claude avec l'outil web_search pour récupérer
    la courbe BAM du jour. Retourne (mats, rats, labels, date_valeur).
    """
    import json

    date_str = eval_date.strftime("%d/%m/%Y")

    prompt = f"""Tu dois récupérer la courbe des taux de référence des bons du Trésor 
publiée par Bank Al-Maghrib (BAM) pour la date du {date_str}.

URL officielle: {_BAM_PAGE_URL}?date={date_str}

Utilise l'outil web_search pour accéder à cette page et extraire le tableau des taux.
Le tableau contient des colonnes: Date d'échéance, Transaction, Taux moyen pondéré, Date de la valeur.

Retourne UNIQUEMENT un JSON valide (sans aucun autre texte) avec cette structure exacte:
{{
  "date_valeur": "JJ/MM/AAAA",
  "points": [
    {{"echeance": "JJ/MM/AAAA", "taux": 2.250}},
    ...
  ]
}}
Les taux sont en pourcentage (ex: 2.250 signifie 2.250%).
"""

    payload = {
        "model": "claude-opus-4-5",
        "max_tokens": 1024,
        "tools": [{"type": "web_search_20250305", "name": "web_search"}],
        "messages": [{"role": "user", "content": prompt}]
    }

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "anthropic-beta": "web-search-2025-03-05",
            "content-type": "application/json",
        },
        json=payload,
        timeout=60,
    )
    resp.raise_for_status()
    data = resp.json()

    # Extract text response
    text = ""
    for block in data.get("content", []):
        if block.get("type") == "text":
            text += block.get("text", "")

    # Parse JSON from response
    text = text.strip()
    # Strip markdown code fences if present
    if "```" in text:
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    text = text.strip()

    result = json.loads(text)
    points = result.get("points", [])
    date_valeur_str = result.get("date_valeur", date_str)

    try:
        date_valeur = datetime.datetime.strptime(date_valeur_str, "%d/%m/%Y").date()
    except Exception:
        date_valeur = eval_date

    mats, rats, labels = [], [], []
    for pt in points:
        try:
            d    = datetime.datetime.strptime(pt["echeance"], "%d/%m/%Y").date()
            rate = float(pt["taux"]) / 100
            days = (d - eval_date).days
            if days <= 0:
                continue
            mats.append(days)
            rats.append(rate)
            labels.append(_days_to_years_label(d, eval_date))
        except Exception:
            continue

    if len(rats) < 2:
        raise ValueError("L'API Claude n'a pas pu extraire suffisamment de points de taux.")

    return mats, rats, labels, date_valeur


def fetch_bam_direct(eval_date: datetime.date):
    """
    Tentative directe via requests (fonctionne en local ou si le réseau le permet).
    """
    date_str = eval_date.strftime("%d/%m/%Y")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,*/*;q=0.8",
        "Accept-Language": "fr-MA,fr;q=0.9",
        "Referer": "https://www.bkam.ma/",
    }
    session = requests.Session()
    session.headers.update(headers)

    urls = [
        f"{_BAM_PAGE_URL}?date={date_str}",
        _BAM_PAGE_URL,
        (
            "https://www.bkam.ma/Marches/Principaux-indicateurs/Marche-obligataire/"
            f"Marche-des-bons-du-tresor/Marche-secondaire/"
            f"Taux-de-reference-des-bons-du-tresor?date={date_str}"
        ),
    ]
    errors = []
    for url in urls:
        try:
            resp = session.get(url, timeout=20, allow_redirects=True)
            if resp.status_code == 200 and len(resp.text) > 200:
                return _parse_bam_html_table(resp.text, eval_date)
            errors.append(f"HTTP {resp.status_code}")
        except Exception as e:
            errors.append(str(e)[:50])
    raise ValueError(" | ".join(errors))



# ─────────────────────────────────────────────
# HELPER: PARSE COURBE MANUELLE
# ─────────────────────────────────────────────
def parse_manual_curve(text: str, eval_date: datetime.date):
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    mats, rats, labels = [], [], []

    mat_map = {
        "13 semaines": 91,  "26 semaines": 182, "52 semaines": 365,
        "1 an": 365,  "2 ans": 730,  "3 ans": 1095,
        "5 ans": 1825, "7 ans": 2555, "10 ans": 3650,
        "15 ans": 5475, "20 ans": 7300, "25 ans": 9125, "30 ans": 10950,
    }

    for line in lines:
        parts = line.replace(",", ".").split()
        if len(parts) < 2:
            continue
        try:
            rate_raw = float(parts[-1].replace("%", ""))
            rate     = rate_raw / 100 if rate_raw > 1 else rate_raw
            key      = " ".join(parts[:-1]).strip().lower()
            days, label = None, None

            # Try DD/MM/YYYY format
            try:
                d     = datetime.datetime.strptime(key.strip(), "%d/%m/%Y").date()
                days  = (d - eval_date).days
                label = _days_to_years_label(d, eval_date)
            except Exception:
                pass

            # Try label match
            if days is None:
                for k, d in mat_map.items():
                    if k.lower() in key or key in k.lower():
                        days, label = d, k
                        break

            # Generic "X ans" / "X semaines"
            if days is None:
                if "an" in key:
                    n = float(key.replace("ans", "").replace("an", "").strip())
                    days, label = int(n * 365), f"{n:.0f} ans"
                elif "sem" in key:
                    n = float(key.replace("semaines", "").replace("semaine", "").strip())
                    days, label = int(n * 7), f"{n:.0f} semaines"

            if days and days > 0:
                mats.append(days)
                rats.append(rate)
                labels.append(label)
        except Exception:
            continue

    if len(rats) < 2:
        raise ValueError("Pas assez de points parsés. Vérifiez le format.")
    return mats, rats, labels

# ─────────────────────────────────────────────
# HELPER: VALORISATION (Circulaire 02/04)
# ─────────────────────────────────────────────
def parse_nominal(val):
    try:
        if isinstance(val, str):
            return float(val.replace("\xa0", "").replace(" ", "").replace(",", "."))
        return float(val) if val else 100000
    except Exception:
        return 100000

def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date):     return v
    return None

def valoriser_titre(row: dict, eval_date: datetime.date, curve_mats: list, curve_rates: list):
    nature  = str(row.get("Nature", "")).strip()
    tf      = float(row.get("Taux facial", 0) or 0)
    spread  = float(row.get("SPREAD", 0) or 0)
    nominal = parse_nominal(row.get("Nominal", 100000))
    nj      = int(row.get("Nb jours", 0) or 0)
    n_coup  = int(row.get("nombre de coupons", 1) or 1)
    A       = int(row.get("A", 365) or 365)

    de  = to_date(row.get("Date Emission"))
    dec = to_date(row.get("Date Échéance"))
    pc  = to_date(row.get("Prochain coupon"))

    if de is None or dec is None:
        return None, None, 0, None, "Données manquantes"

    Mi = (dec - de).days
    Mr = (dec - eval_date).days

    if Mr <= 0:
        return None, None, 0, None, "Titre échu"

    tr_bam = interpolate_rate(Mr, curve_mats, curve_rates)
    tr     = tr_bam + spread

    price, cc, methode = None, 0.0, ""

    # Formule (1): Mi ≤ 365
    if Mi <= 365:
        methode = "Formule (1) Mi≤365j"
        price   = nominal * (1 + tf * Mi / 360) / (1 + tr * Mr / 360)

    # Formule (2) ou (3): Mi > 365, Mr ≤ 365
    elif Mr <= 365:
        is_post_1flux = ("Atypique" in nature or "postérieure" in nature.lower()) and n_coup <= 1
        if is_post_1flux:
            methode = "Formule (3) Ligne postérieure 1 flux"
            price   = nominal * (1 + tf * Mi / A) / (1 + tr * Mr / 360)
        else:
            methode = "Formule (2) Mi>365 & Mr≤365"
            price   = nominal * (1 + tf) / (1 + tr * Mr / 360)
        if pc:
            days_to_next = (pc - eval_date).days
            if 0 < days_to_next <= A:
                cc = nominal * tf * (1 - days_to_next / A)

    # Formule (4): Mi > 365, Mr > 365
    else:
        methode = "Formule (4) Mi>365 & Mr>365"
        price   = 0.0
        for i in range(n_coup):
            Fi  = nominal * tf if i < n_coup - 1 else nominal * (1 + tf)
            exp = nj / A + i
            price += Fi / (1 + tr) ** exp
        if pc and nj > 0:
            cc = nominal * tf * (1 - nj / A)

    if price is None:
        return None, None, 0, tr, methode

    ppc = price - cc
    return round(price, 4), round(ppc, 4), round(cc, 4), tr, methode

# ─────────────────────────────────────────────
# HELPER: LOAD EXCEL
# ─────────────────────────────────────────────
def load_excel(file):
    wb   = load_workbook(file, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    headers = rows[0]
    data    = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
    return pd.DataFrame(data)


# ═══════════════════════════════════════════════════════════
# PAGE 1 – ACCUEIL & COURBE BAM
# ═══════════════════════════════════════════════════════════
if "Accueil" in page:

    col_main, col_side = st.columns([3, 1], gap="large")

    with col_main:
        st.markdown('<div class="section-title">📡 Importation de la Courbe des Taux BAM</div>', unsafe_allow_html=True)

        tab_csv, tab_manual = st.tabs([
            "📁 Import CSV BAM",
            "✏️ Saisie Manuelle",
        ])

        # ── Tab 1 : Import CSV officiel ───────────────────────────────
        with tab_csv:
            # Lien direct + instructions visuelles en haut
            bam_page_url = (
                "https://www.bkam.ma/Marches/Principaux-indicateurs/Marche-obligataire/"
                "Marche-des-bons-de-tresor/Marche-secondaire/"
                "Taux-de-reference-des-bons-du-tresor"
            )
            st.markdown(f"""
            <div style="background:linear-gradient(90deg,#1a1a2e,#0f3460);
                        border-radius:10px;padding:16px 20px;margin-bottom:16px;">
                <div style="color:#fff;font-weight:700;font-size:.95rem;margin-bottom:10px;">
                    📋 Procédure en 3 étapes
                </div>
                <div style="display:flex;gap:12px;flex-wrap:wrap;">
                    <div style="background:rgba(255,255,255,.08);border-radius:8px;
                                padding:12px 16px;flex:1;min-width:160px;">
                        <div style="color:#e31e24;font-weight:700;font-size:1.1rem;">① Ouvrir</div>
                        <div style="color:#ccc;font-size:.82rem;margin-top:4px;">
                            Cliquez sur le bouton ci-dessous pour aller sur bkam.ma
                        </div>
                    </div>
                    <div style="background:rgba(255,255,255,.08);border-radius:8px;
                                padding:12px 16px;flex:1;min-width:160px;">
                        <div style="color:#e31e24;font-weight:700;font-size:1.1rem;">② Télécharger</div>
                        <div style="color:#ccc;font-size:.82rem;margin-top:4px;">
                            Choisissez la date → cliquez <b style="color:#fff">Téléchargement CSV</b>
                        </div>
                    </div>
                    <div style="background:rgba(255,255,255,.08);border-radius:8px;
                                padding:12px 16px;flex:1;min-width:160px;">
                        <div style="color:#e31e24;font-weight:700;font-size:1.1rem;">③ Importer</div>
                        <div style="color:#ccc;font-size:.82rem;margin-top:4px;">
                            Glissez le fichier CSV dans la zone ci-dessous
                        </div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            col_btn, col_date = st.columns([2, 1])
            with col_btn:
                st.link_button(
                    "🌐 Ouvrir bkam.ma – Taux de référence BDT",
                    url=bam_page_url,
                    use_container_width=True,
                )
            with col_date:
                eval_date_csv = st.date_input(
                    "Date d'évaluation",
                    value=datetime.date.today(),
                    format="DD/MM/YYYY",
                    key="csv_eval_date",
                )

            uploaded_csv = st.file_uploader(
                "📂 Déposez ici le fichier CSV téléchargé depuis bkam.ma",
                type=["csv", "txt"],
                key="bam_csv_upload",
                help="Fichier nommé : Taux-de-reference-des-bons-du-tresor_JJ_MM_AAAA.csv",
            )

            if uploaded_csv is not None:
                raw = uploaded_csv.read()
                if st.button("📥 Charger cette courbe", key="btn_csv_load", use_container_width=True):
                    try:
                        mats, rats, labels, dv = parse_bam_csv(raw, eval_date_csv)
                        st.session_state.bam_curve = dict(
                            maturities_days=mats, rates=rats, labels=labels, date=dv
                        )
                        st.success(
                            f"✅ Courbe chargée — **{len(rats)} points** — "
                            f"Date valeur : **{dv.strftime('%d/%m/%Y')}**"
                        )
                        df_prev = pd.DataFrame({
                            "Échéance": labels,
                            "Taux (%)": [f"{r*100:.4f}%" for r in rats],
                        })
                        st.dataframe(df_prev, use_container_width=True, hide_index=True)
                    except Exception as e:
                        st.error(f"❌ Erreur : {e}")

        # ── Tab 2 : Saisie manuelle ───────────────────────────────────
        with tab_manual:
            st.markdown("**Saisie manuelle de la courbe des taux**")
            st.caption(
                "Formats acceptés (une ligne par point) :\n"
                "- `DD/MM/YYYY  taux` → ex : `18/05/2026  2.250`\n"
                "- `X ans  taux`      → ex : `10 ans  3.460`\n"
                "- `X semaines  taux` → ex : `13 semaines  2.250`\n"
                "_Le taux peut être en % (ex: 2.25) ou en décimal (ex: 0.0225)._"
            )
            col1, col2 = st.columns([3, 1])
            with col1:
                default_manual = (
                    "18/05/2026  2.250\n"
                    "17/08/2026  2.290\n"
                    "15/02/2027  2.390\n"
                    "20/09/2027  2.440\n"
                    "15/05/2028  2.750\n"
                    "20/10/2031  2.900\n"
                    "18/06/2035  3.140\n"
                    "18/07/2039  3.460\n"
                    "14/08/2045  3.700\n"
                    "19/04/2055  4.000"
                )
                manual_txt = st.text_area(
                    "Courbe des taux",
                    value=default_manual,
                    height=280,
                    help="Données du 16/04/2026 à titre d'exemple",
                )
            with col2:
                eval_date_man = st.date_input(
                    "Date d'évaluation",
                    value=datetime.date.today(),
                    format="DD/MM/YYYY",
                    key="man_eval_date",
                )
                st.markdown("<br/>", unsafe_allow_html=True)
                if st.button("✅ Valider", key="btn_manual"):
                    try:
                        mats, rats, labels = parse_manual_curve(manual_txt, eval_date_man)
                        st.session_state.bam_curve = dict(
                            maturities_days=mats, rates=rats, labels=labels, date=eval_date_man
                        )
                        st.success(f"✅ {len(rats)} points enregistrés")
                    except Exception as e:
                        st.error(f"❌ {e}")

    # ── Panneau latéral : courbe chargée ─────────────────────────────
    with col_side:
        st.markdown('<div class="section-title">📊 Courbe chargée</div>', unsafe_allow_html=True)
        curve = st.session_state.bam_curve
        if curve:
            st.success(
                f"✅ **{len(curve['rates'])} points**\n\n"
                f"Date : **{curve['date'].strftime('%d/%m/%Y')}**"
            )
            fig_mini = go.Figure()
            fig_mini.add_trace(go.Scatter(
                x=curve["labels"],
                y=[r * 100 for r in curve["rates"]],
                mode="lines+markers",
                line=dict(color=CFG_RED, width=2.5),
                marker=dict(size=7, color=CFG_DARK),
                fill="tozeroy",
                fillcolor="rgba(227,30,36,0.08)",
            ))
            fig_mini.update_layout(
                height=260,
                margin=dict(l=30, r=10, t=10, b=70),
                plot_bgcolor="white", paper_bgcolor="white", showlegend=False,
                xaxis=dict(tickangle=-45, tickfont=dict(size=8), showgrid=True, gridcolor="#eee"),
                yaxis=dict(title="Taux (%)", tickfont=dict(size=9), showgrid=True, gridcolor="#eee"),
            )
            st.plotly_chart(fig_mini, use_container_width=True)

            df_c = pd.DataFrame({
                "Échéance": curve["labels"],
                "Taux (%)": [f"{r*100:.4f}%" for r in curve["rates"]],
            })
            st.dataframe(df_c, use_container_width=True, hide_index=True)
            if st.button("🗑️ Effacer", key="btn_clear"):
                st.session_state.bam_curve = None
                st.rerun()
        else:
            st.info("Aucune courbe chargée.")

    # ── Chargement fichier Excel ──────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">📁 Fichier de Données (DATA_VALORISATION.xlsx)</div>', unsafe_allow_html=True)

    col_up, col_stat = st.columns([3, 1], gap="large")
    with col_up:
        uploaded_xl = st.file_uploader(
            "Glissez-déposez votre fichier Excel",
            type=["xlsx", "xls"],
            help="Colonnes requises : Nature, TITRE, CODE ISIN, Date Emission, Date Échéance, "
                 "Taux facial, SPREAD, Nominal, Nb jours, nombre de coupons, A, Prochain coupon",
        )
        if uploaded_xl is not None:
            with st.spinner("Lecture du fichier…"):
                df = load_excel(uploaded_xl)
            if df is not None and not df.empty:
                st.session_state.df_titres = df
                st.success(f"✅ **{len(df)} titres** importés")
                st.dataframe(
                    df[["TITRE", "CODE ISIN", "Nature", "Date Emission",
                        "Date Échéance", "Taux facial", "SPREAD", "Nominal"]].head(8),
                    use_container_width=True, hide_index=True,
                )
            else:
                st.error("Fichier vide ou illisible.")

    with col_stat:
        df_t = st.session_state.df_titres
        if df_t is not None:
            st.metric("Titres chargés", len(df_t))
            natures = df_t["Nature"].value_counts()
            fig_d = go.Figure(go.Pie(
                labels=[str(n)[:28] for n in natures.index],
                values=natures.values,
                hole=0.55,
                marker_colors=COLORS,
            ))
            fig_d.update_layout(
                height=200, margin=dict(l=0, r=0, t=5, b=0),
                showlegend=False, paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_d, use_container_width=True)


# ═══════════════════════════════════════════════════════════
# PAGE 2 – VALORISATION
# ═══════════════════════════════════════════════════════════
elif "Valorisation" in page:

    if st.session_state.df_titres is None:
        st.warning("⚠️ Chargez d'abord votre fichier dans **Accueil & Courbe BAM**.")
        st.stop()
    if st.session_state.bam_curve is None:
        st.warning("⚠️ Importez d'abord la courbe BAM dans **Accueil & Courbe BAM**.")
        st.stop()

    st.markdown('<div class="section-title">⚙️ Paramètres de Valorisation</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        eval_date = st.date_input("📅 Date d'évaluation", value=datetime.date.today(), format="DD/MM/YYYY")
    with col2:
        df_t = st.session_state.df_titres
        filter_nature = st.multiselect(
            "Filtrer par Nature",
            options=["Tous"] + sorted(df_t["Nature"].dropna().unique().tolist()),
            default=["Tous"],
        )
    with col3:
        modeles = sorted(df_t["MODELE"].dropna().unique().tolist())
        filter_modele = st.multiselect(
            "Filtrer par Modèle",
            options=["Tous"] + modeles,
            default=["Tous"],
        )

    df_work = df_t.copy()
    if "Tous" not in filter_nature and filter_nature:
        df_work = df_work[df_work["Nature"].isin(filter_nature)]
    if "Tous" not in filter_modele and filter_modele:
        df_work = df_work[df_work["MODELE"].isin(filter_modele)]

    st.info(f"🔢 **{len(df_work)} titres** sélectionnés pour la valorisation")

    if st.button("🚀 Lancer la Valorisation", key="btn_run"):
        curve = st.session_state.bam_curve
        results = []
        prog  = st.progress(0, text="Valorisation en cours…")
        total = len(df_work)

        for i, (_, row) in enumerate(df_work.iterrows()):
            r = row.to_dict()
            price, ppc, cc, tr, methode = valoriser_titre(
                r, eval_date, curve["maturities_days"], curve["rates"]
            )
            nom = parse_nominal(r.get("Nominal", 100000))
            de  = to_date(r.get("Date Emission"))
            dec = to_date(r.get("Date Échéance"))
            Mi  = (dec - de).days   if dec and de        else None
            Mr  = (dec - eval_date).days if dec          else None

            results.append({
                "TITRE":            r.get("TITRE", ""),
                "CODE ISIN":        r.get("CODE ISIN", ""),
                "Nature":           r.get("Nature", ""),
                "MODELE":           r.get("MODELE", ""),
                "Date Emission":    de,
                "Date Échéance":    dec,
                "Mi (j)":           Mi,
                "Mr (j)":           Mr,
                "Taux Facial (%)":  float(r.get("Taux facial", 0) or 0) * 100,
                "Spread (%)":       float(r.get("SPREAD", 0) or 0) * 100,
                "Taux Act. (%)":    round(tr * 100, 6) if tr else None,
                "Nominal":          nom,
                "Prix (%)":         round(price / nom * 100, 6) if price and nom else None,
                "Prix Absolu":      price,
                "Coupon Couru":     cc,
                "Prix Pied Coupon": ppc,
                "Méthode":          methode,
                "Statut":           "✅ OK" if price is not None else "❌ Erreur",
            })
            prog.progress((i + 1) / total,
                          text=f"Traitement {i+1}/{total} – {r.get('TITRE','')[:35]}")

        prog.empty()
        df_res = pd.DataFrame(results)
        st.session_state.df_results = df_res

        ok = df_res[df_res["Statut"] == "✅ OK"]
        st.success(f"✅ Terminé — {len(ok)}/{total} titres valorisés avec succès")

        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi-card">
                <div class="kpi-label">Titres OK</div>
                <div class="kpi-value">{len(ok)}</div>
                <div class="kpi-sub">sur {total}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Valeur Totale</div>
                <div class="kpi-value">{ok['Prix Absolu'].sum():,.0f}</div>
                <div class="kpi-sub">MAD</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Prix Moyen</div>
                <div class="kpi-value">{ok['Prix (%)'].mean():.4f}%</div>
                <div class="kpi-sub">du nominal</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Taux Act. Moyen</div>
                <div class="kpi-value">{ok['Taux Act. (%)'].mean():.4f}%</div>
                <div class="kpi-sub">interpolé</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.dataframe(df_res, use_container_width=True, hide_index=True)

    elif st.session_state.df_results is not None:
        df_res = st.session_state.df_results
        ok = df_res[df_res["Statut"] == "✅ OK"]
        st.info("Résultats du dernier calcul (relancez pour recalculer) :")
        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi-card"><div class="kpi-label">Titres OK</div>
                <div class="kpi-value">{len(ok)}</div><div class="kpi-sub">sur {len(df_res)}</div></div>
            <div class="kpi-card"><div class="kpi-label">Valeur Totale</div>
                <div class="kpi-value">{ok['Prix Absolu'].sum():,.0f}</div><div class="kpi-sub">MAD</div></div>
            <div class="kpi-card"><div class="kpi-label">Prix Moyen</div>
                <div class="kpi-value">{ok['Prix (%)'].mean():.4f}%</div><div class="kpi-sub">du nominal</div></div>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(df_res, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════
# PAGE 3 – ANALYSE & GRAPHIQUES
# ═══════════════════════════════════════════════════════════
elif "Analyse" in page:

    if st.session_state.df_results is None:
        st.warning("⚠️ Lancez d'abord la valorisation dans l'onglet **Valorisation**.")
        st.stop()

    df_res = st.session_state.df_results.copy()
    ok = df_res[df_res["Statut"] == "✅ OK"].copy()

    if ok.empty:
        st.error("Aucun titre valorisé avec succès.")
        st.stop()

    # ── 1. Courbe BAM ─────────────────────────────────────────────────
    if st.session_state.bam_curve:
        st.markdown('<div class="section-title">📉 Courbe des Taux de Référence BAM</div>', unsafe_allow_html=True)
        curve  = st.session_state.bam_curve
        x_lab  = curve["labels"]
        y_val  = [r * 100 for r in curve["rates"]]

        fig_bam = go.Figure()
        fig_bam.add_trace(go.Scatter(
            x=x_lab, y=y_val,
            mode="lines+markers+text",
            line=dict(color=CFG_RED, width=3),
            marker=dict(size=10, color=CFG_DARK, line=dict(color=CFG_RED, width=2)),
            text=[f"{v:.3f}%" for v in y_val],
            textposition="top center",
            textfont=dict(size=10, color=CFG_DARK),
            fill="tozeroy",
            fillcolor="rgba(227,30,36,0.07)",
            name="Taux BAM",
        ))
        fig_bam = chart_layout(
            fig_bam,
            f"Courbe des Taux de Référence BAM – Date valeur : {curve['date'].strftime('%d/%m/%Y')}",
            h=380, legend=False,
        )
        fig_bam.update_xaxes(title_text="Échéance", tickangle=-30, tickfont=dict(size=10))
        fig_bam.update_yaxes(title_text="Taux (%)", tickformat=".3f")
        st.plotly_chart(fig_bam, use_container_width=True)

    st.markdown("---")

    # ── 2. Répartition par formule + Boxplot prix ─────────────────────
    st.markdown('<div class="section-title">📊 Répartition & Statistiques</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2, gap="large")

    with col1:
        mc = ok["Méthode"].value_counts().reset_index()
        mc.columns = ["Méthode", "Nb"]
        fig_meth = go.Figure(go.Bar(
            x=mc["Nb"],
            y=mc["Méthode"],
            orientation="h",
            marker=dict(
                color=COLORS[:len(mc)],
                line=dict(color="white", width=0.8),
            ),
            text=mc["Nb"],
            textposition="outside",
            textfont=dict(size=12, color=CFG_DARK),
        ))
        fig_meth = chart_layout(fig_meth, "Titres par Formule de Valorisation", h=340, legend=False)
        fig_meth.update_xaxes(title_text="Nombre de titres")
        fig_meth.update_yaxes(tickfont=dict(size=10), autorange="reversed")
        st.plotly_chart(fig_meth, use_container_width=True)

    with col2:
        fig_box = go.Figure()
        fig_box.add_trace(go.Box(
            y=ok["Prix (%)"],
            name="Prix (%)",
            boxpoints="outliers",
            marker=dict(color=CFG_RED, size=5),
            line=dict(color=CFG_DARK),
            fillcolor="rgba(227,30,36,0.12)",
            whiskerwidth=0.7,
        ))
        fig_box = chart_layout(fig_box, "Distribution des Prix (% du Nominal)", h=340, legend=False)
        fig_box.update_yaxes(title_text="Prix (%)")
        # Annotate quartiles
        q1  = ok["Prix (%)"].quantile(0.25)
        med = ok["Prix (%)"].median()
        q3  = ok["Prix (%)"].quantile(0.75)
        for val, lbl, clr in [(q1, f"Q1:{q1:.2f}%", CFG_BLUE),
                               (med, f"Med:{med:.2f}%", CFG_RED),
                               (q3, f"Q3:{q3:.2f}%", CFG_BLUE)]:
            fig_box.add_hline(y=val, line_dash="dot", line_color=clr, line_width=1,
                              annotation_text=lbl, annotation_position="right",
                              annotation_font=dict(size=10, color=clr))
        st.plotly_chart(fig_box, use_container_width=True)

    # ── 3. Histogrammes ───────────────────────────────────────────────
    st.markdown('<div class="section-title">📈 Distributions</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2, gap="large")

    with col1:
        mean_p = ok["Prix (%)"].mean()
        med_p  = ok["Prix (%)"].median()
        fig_h1 = go.Figure()
        fig_h1.add_trace(go.Histogram(
            x=ok["Prix (%)"], nbinsx=25, name="Prix (%)",
            marker=dict(color=CFG_RED, opacity=0.8, line=dict(color="white", width=0.5)),
        ))
        fig_h1.add_vline(x=mean_p, line_dash="dash", line_color=CFG_DARK, line_width=2,
                         annotation_text=f" Moy: {mean_p:.2f}%",
                         annotation_font=dict(size=11, color=CFG_DARK),
                         annotation_position="top right")
        fig_h1.add_vline(x=med_p, line_dash="dot", line_color=CFG_BLUE, line_width=2,
                         annotation_text=f" Méd: {med_p:.2f}%",
                         annotation_font=dict(size=11, color=CFG_BLUE),
                         annotation_position="top left")
        fig_h1 = chart_layout(fig_h1, "Distribution des Prix (% du Nominal)", h=360, legend=False)
        fig_h1.update_xaxes(title_text="Prix (%)")
        fig_h1.update_yaxes(title_text="Nombre de titres")
        st.plotly_chart(fig_h1, use_container_width=True)

    with col2:
        mean_t = ok["Taux Act. (%)"].mean()
        fig_h2 = go.Figure()
        fig_h2.add_trace(go.Histogram(
            x=ok["Taux Act. (%)"], nbinsx=25, name="Taux Act. (%)",
            marker=dict(color=CFG_BLUE, opacity=0.8, line=dict(color="white", width=0.5)),
        ))
        fig_h2.add_vline(x=mean_t, line_dash="dash", line_color=CFG_RED, line_width=2,
                         annotation_text=f" Moy: {mean_t:.3f}%",
                         annotation_font=dict(size=11, color=CFG_RED),
                         annotation_position="top right")
        fig_h2 = chart_layout(fig_h2, "Distribution des Taux d'Actualisation", h=360, legend=False)
        fig_h2.update_xaxes(title_text="Taux d'Actualisation (%)")
        fig_h2.update_yaxes(title_text="Nombre de titres")
        st.plotly_chart(fig_h2, use_container_width=True)

    # ── 4. Scatter Taux Act. vs Maturité ─────────────────────────────
    st.markdown('<div class="section-title">🔵 Taux d\'Actualisation vs Maturité Résiduelle</div>', unsafe_allow_html=True)

    ok_sc = ok.dropna(subset=["Mr (j)", "Taux Act. (%)", "Prix (%)"]).copy()
    ok_sc["Mr (ans)"] = (ok_sc["Mr (j)"] / 365).round(2)

    fig_sc = go.Figure()
    color_iter = iter(COLORS)
    for met, grp in ok_sc.groupby("Méthode"):
        c = next(color_iter, CFG_DARK)
        fig_sc.add_trace(go.Scatter(
            x=grp["Mr (ans)"],
            y=grp["Taux Act. (%)"],
            mode="markers",
            name=met,
            marker=dict(size=10, color=c, opacity=0.8,
                        line=dict(width=0.5, color="white")),
            text=grp["TITRE"],
            customdata=grp[["Prix (%)", "Spread (%)"]].values,
            hovertemplate=(
                "<b>%{text}</b><br>"
                "Maturité: %{x:.2f} ans<br>"
                "Taux Act.: %{y:.4f}%<br>"
                "Prix: %{customdata[0]:.4f}%<br>"
                "Spread: %{customdata[1]:.4f}%<extra></extra>"
            ),
        ))

    # Overlay BAM curve
    if st.session_state.bam_curve:
        c = st.session_state.bam_curve
        fig_sc.add_trace(go.Scatter(
            x=[d / 365 for d in c["maturities_days"]],
            y=[r * 100 for r in c["rates"]],
            mode="lines+markers",
            name="Courbe BAM",
            line=dict(color="black", width=2, dash="dash"),
            marker=dict(size=6, color="black"),
        ))

    fig_sc = chart_layout(
        fig_sc,
        "Taux d'Actualisation vs Maturité Résiduelle — Courbe BAM en tirets",
        h=460,
    )
    fig_sc.update_xaxes(title_text="Maturité Résiduelle (années)")
    fig_sc.update_yaxes(title_text="Taux d'Actualisation (%)")
    st.plotly_chart(fig_sc, use_container_width=True)

    # ── 5. Top / Bottom Prix ──────────────────────────────────────────
    st.markdown('<div class="section-title">🏆 Classement des Titres</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2, gap="large")

    def bar_h(df_in, col_x, title, color_scale):
        df_in = df_in.copy()
        df_in["Label"] = df_in["TITRE"].str[:30]
        fig = go.Figure(go.Bar(
            x=df_in[col_x],
            y=df_in["Label"],
            orientation="h",
            marker=dict(
                color=df_in[col_x],
                colorscale=color_scale,
                showscale=False,
                line=dict(color="white", width=0.5),
            ),
            text=[f"{v:.4f}%" for v in df_in[col_x]],
            textposition="outside",
            textfont=dict(size=10),
        ))
        fig = chart_layout(fig, title, h=400, legend=False)
        fig.update_xaxes(title_text="Prix (%)")
        fig.update_yaxes(autorange="reversed", tickfont=dict(size=10))
        fig.update_layout(margin=dict(l=200, r=80, t=55, b=50))
        return fig

    with col1:
        top10 = ok.nlargest(10, "Prix (%)")[["TITRE", "Prix (%)", "Mr (j)"]].reset_index(drop=True)
        st.plotly_chart(
            bar_h(top10, "Prix (%)", "Top 10 – Prix les Plus Élevés",
                  [[0, "#f7b6b8"], [1, CFG_RED]]),
            use_container_width=True,
        )

    with col2:
        bot10 = ok.nsmallest(10, "Prix (%)")[["TITRE", "Prix (%)", "Mr (j)"]].reset_index(drop=True)
        st.plotly_chart(
            bar_h(bot10, "Prix (%)", "Top 10 – Prix les Plus Faibles",
                  [[0, CFG_BLUE], [1, "#a0b4d0"]]),
            use_container_width=True,
        )

    # ── 6. Spread vs Prix (bubble) ────────────────────────────────────
    st.markdown('<div class="section-title">📌 Spread vs Prix — Taille des bulles ∝ Maturité</div>', unsafe_allow_html=True)

    ok_sp = ok.dropna(subset=["Spread (%)", "Prix (%)"]).copy()
    ok_sp["Mr (ans)"] = (ok_sp["Mr (j)"].fillna(365) / 365).round(1).clip(1, 35)

    fig_sp = go.Figure()
    color_iter2 = iter(COLORS)
    for met, grp in ok_sp.groupby("Méthode"):
        c = next(color_iter2, CFG_DARK)
        fig_sp.add_trace(go.Scatter(
            x=grp["Spread (%)"],
            y=grp["Prix (%)"],
            mode="markers",
            name=met,
            marker=dict(
                size=grp["Mr (ans)"] * 1.5 + 5,
                color=c,
                opacity=0.72,
                line=dict(width=0.5, color="white"),
            ),
            text=grp["TITRE"],
            customdata=grp["Mr (ans)"].values,
            hovertemplate=(
                "<b>%{text}</b><br>"
                "Spread: %{x:.4f}%<br>"
                "Prix: %{y:.4f}%<br>"
                "Maturité: %{customdata:.1f} ans<extra></extra>"
            ),
        ))

    fig_sp = chart_layout(fig_sp, "Spread vs Prix (bulle = maturité résiduelle)", h=460)
    fig_sp.update_xaxes(title_text="Spread (%)")
    fig_sp.update_yaxes(title_text="Prix (%)")
    st.plotly_chart(fig_sp, use_container_width=True)

    # ── 7. Coupons courus ─────────────────────────────────────────────
    ok_cc = ok[ok["Coupon Couru"] > 0].nlargest(15, "Coupon Couru").copy()
    if not ok_cc.empty:
        st.markdown('<div class="section-title">💰 Coupons Courus (Top 15)</div>', unsafe_allow_html=True)
        ok_cc["Label"] = ok_cc["TITRE"].str[:28]
        fig_cc = go.Figure(go.Bar(
            x=ok_cc["Label"],
            y=ok_cc["Coupon Couru"],
            marker=dict(color=CFG_RED, opacity=0.85, line=dict(color="white", width=0.5)),
            text=[f"{v:,.0f}" for v in ok_cc["Coupon Couru"]],
            textposition="outside",
            textfont=dict(size=10),
        ))
        fig_cc = chart_layout(fig_cc, "Coupon Couru par Titre – Top 15 (MAD)", h=380, legend=False)
        fig_cc.update_xaxes(tickangle=-40, tickfont=dict(size=9))
        fig_cc.update_yaxes(title_text="Coupon Couru (MAD)")
        st.plotly_chart(fig_cc, use_container_width=True)


# ═══════════════════════════════════════════════════════════
# PAGE 4 – RÉSULTATS & EXPORT
# ═══════════════════════════════════════════════════════════
elif "Résultats" in page:

    if st.session_state.df_results is None:
        st.warning("⚠️ Lancez d'abord la valorisation.")
        st.stop()

    df_res = st.session_state.df_results.copy()
    ok     = df_res[df_res["Statut"] == "✅ OK"]

    st.markdown('<div class="section-title">📋 Tableau des Résultats</div>', unsafe_allow_html=True)

    # Filtres
    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
    with col1:
        search = st.text_input("🔍 Recherche (TITRE ou ISIN)", "")
    with col2:
        statut_f = st.selectbox("Statut", ["Tous", "✅ OK", "❌ Erreur"])
    with col3:
        sort_col = st.selectbox("Trier par", ["TITRE", "Prix (%)", "Taux Act. (%)", "Mr (j)", "Spread (%)"])
    with col4:
        sort_asc = st.radio("Ordre", ["↓ Desc", "↑ Asc"], horizontal=True) == "↑ Asc"

    df_disp = df_res.copy()
    if search:
        mask = (
            df_disp["TITRE"].str.contains(search, case=False, na=False) |
            df_disp["CODE ISIN"].str.contains(search, case=False, na=False)
        )
        df_disp = df_disp[mask]
    if statut_f == "✅ OK":
        df_disp = df_disp[df_disp["Statut"] == "✅ OK"]
    elif statut_f == "❌ Erreur":
        df_disp = df_disp[df_disp["Statut"] != "✅ OK"]
    df_disp = df_disp.sort_values(sort_col, ascending=sort_asc, na_position="last")

    ok_d = df_disp[df_disp["Statut"] == "✅ OK"]
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Affichés", len(df_disp))
    c2.metric("Valorisés ✅", len(ok_d))
    c3.metric("Erreurs ❌", len(df_disp) - len(ok_d))
    c4.metric("Valeur Totale", f"{ok_d['Prix Absolu'].sum():,.0f} MAD" if len(ok_d) else "—")
    c5.metric("Prix Moyen",   f"{ok_d['Prix (%)'].mean():.4f}%"       if len(ok_d) else "—")

    st.markdown("---")

    num_fmt = {
        "Taux Facial (%)":  "{:.4f}%",
        "Spread (%)":       "{:.4f}%",
        "Taux Act. (%)":    "{:.4f}%",
        "Prix (%)":         "{:.6f}%",
        "Prix Absolu":      "{:,.4f}",
        "Coupon Couru":     "{:,.4f}",
        "Prix Pied Coupon": "{:,.4f}",
        "Nominal":          "{:,.2f}",
    }

    st.dataframe(
        df_disp.style
            .apply(lambda r: ["background-color:#fff5f5" if r["Statut"] != "✅ OK" else "" for _ in r], axis=1)
            .format(num_fmt, na_rep="—"),
        use_container_width=True,
        height=480,
    )

    # Détail titre
    st.markdown('<div class="section-title">🔎 Détail par Titre</div>', unsafe_allow_html=True)
    titres_list = df_disp["TITRE"].tolist()
    if titres_list:
        titre_sel = st.selectbox("Sélectionner un titre", titres_list)
        r = df_disp[df_disp["TITRE"] == titre_sel].iloc[0]
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Prix (%)",          f"{r['Prix (%)']:.6f}%"       if r['Prix (%)']         else "—")
        col2.metric("Prix Absolu (MAD)", f"{r['Prix Absolu']:,.4f}"     if r['Prix Absolu']       else "—")
        col3.metric("Coupon Couru",      f"{r['Coupon Couru']:,.4f}"    if r['Coupon Couru']      else "0")
        col4.metric("Pied de Coupon",    f"{r['Prix Pied Coupon']:,.4f}" if r['Prix Pied Coupon'] else "—")
        col5, col6, col7, col8 = st.columns(4)
        col5.metric("Méthode",     r["Méthode"])
        col6.metric("Mr (jours)", str(r["Mr (j)"]) if r["Mr (j)"] else "—")
        col7.metric("Taux Facial", f"{r['Taux Facial (%)']:.4f}%")
        col8.metric("Taux Act.",   f"{r['Taux Act. (%)']:.4f}%" if r['Taux Act. (%)'] else "—")

    # Export
    st.markdown("---")
    st.markdown('<div class="section-title">💾 Export des Résultats</div>', unsafe_allow_html=True)
    col_e1, col_e2, col_e3 = st.columns(3)

    with col_e1:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_res.to_excel(writer, sheet_name="Valorisation", index=False)
            if st.session_state.bam_curve:
                c = st.session_state.bam_curve
                pd.DataFrame({
                    "Échéance": c["labels"],
                    "Jours": c["maturities_days"],
                    "Taux (%)": [r * 100 for r in c["rates"]],
                }).to_excel(writer, sheet_name="Courbe BAM", index=False)
            if len(ok) > 0:
                pd.DataFrame({
                    "Indicateur": [
                        "Titres valorisés", "Valeur Totale (MAD)",
                        "Prix Moyen (%)", "Taux Act. Moyen (%)", "Date évaluation",
                    ],
                    "Valeur": [
                        len(ok),
                        f"{ok['Prix Absolu'].sum():,.2f}",
                        f"{ok['Prix (%)'].mean():.6f}%",
                        f"{ok['Taux Act. (%)'].mean():.4f}%",
                        datetime.date.today().strftime("%d/%m/%Y"),
                    ],
                }).to_excel(writer, sheet_name="Résumé", index=False)
        st.download_button(
            "📥 Excel Complet (3 onglets)",
            data=buf.getvalue(),
            file_name=f"Valorisation_OPCVM_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col_e2:
        csv_buf = io.StringIO()
        df_res.to_csv(csv_buf, index=False, sep=";", decimal=",", encoding="utf-8-sig")
        st.download_button(
            "📥 CSV (séparateur ;)",
            data=csv_buf.getvalue().encode("utf-8-sig"),
            file_name=f"Valorisation_OPCVM_{datetime.date.today().strftime('%Y%m%d')}.csv",
            mime="text/csv",
        )

    with col_e3:
        buf2 = io.BytesIO()
        with pd.ExcelWriter(buf2, engine="openpyxl") as writer:
            ok.to_excel(writer, sheet_name="Valorisation OK", index=False)
        st.download_button(
            "📥 Excel – Titres OK uniquement",
            data=buf2.getvalue(),
            file_name=f"Valorisation_OK_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    errors = df_res[df_res["Statut"] != "✅ OK"]
    if not errors.empty:
        with st.expander(f"⚠️ Titres en erreur ({len(errors)})", expanded=False):
            st.dataframe(
                errors[["TITRE", "CODE ISIN", "Nature", "Méthode", "Statut"]],
                use_container_width=True, hide_index=True,
            )
