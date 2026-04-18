import streamlit as st
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import io
import datetime
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import base64
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="CFG Bank – Valorisation OPCVM",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* ── Global ──────────────────────────────── */
    [data-testid="stAppViewContainer"] { background: #f4f6fa; }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
    }
    [data-testid="stSidebar"] * { color: #e0e0e0 !important; }
    [data-testid="stSidebar"] hr { border-color: #e31e24 !important; }

    /* ── Top header bar ──────────────────────── */
    .cfg-header {
        background: linear-gradient(90deg, #1a1a2e 0%, #e31e24 100%);
        padding: 18px 30px;
        border-radius: 12px;
        display: flex;
        align-items: center;
        gap: 24px;
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(227,30,36,.25);
    }
    .cfg-header h1 { color: #fff; margin: 0; font-size: 1.55rem; font-weight: 700; letter-spacing: .5px; }
    .cfg-header p  { color: #ffd0d0; margin: 4px 0 0; font-size: .85rem; }

    /* ── KPI cards ───────────────────────────── */
    .kpi-row { display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }
    .kpi-card {
        background: #fff;
        border-radius: 12px;
        padding: 18px 22px;
        flex: 1;
        min-width: 160px;
        box-shadow: 0 2px 12px rgba(0,0,0,.07);
        border-top: 4px solid #e31e24;
    }
    .kpi-card .kpi-label { font-size: .78rem; color: #888; text-transform: uppercase; letter-spacing: .5px; }
    .kpi-card .kpi-value { font-size: 1.5rem; font-weight: 700; color: #1a1a2e; margin-top: 4px; }
    .kpi-card .kpi-sub   { font-size: .78rem; color: #aaa; margin-top: 2px; }

    /* ── Section titles ──────────────────────── */
    .section-title {
        font-size: 1.05rem; font-weight: 700; color: #1a1a2e;
        border-left: 4px solid #e31e24; padding-left: 10px;
        margin: 20px 0 12px;
    }

    /* ── Tables ──────────────────────────────── */
    .dataframe thead th { background: #1a1a2e !important; color: #fff !important; font-size: .82rem; }
    .dataframe tbody tr:nth-child(even) { background: #f9f0f0 !important; }
    .dataframe tbody td { font-size: .82rem; }

    /* ── Buttons ─────────────────────────────── */
    .stButton > button {
        background: linear-gradient(90deg, #e31e24, #c0141a);
        color: #fff !important;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        padding: 10px 22px;
        box-shadow: 0 3px 10px rgba(227,30,36,.3);
        transition: .2s;
    }
    .stButton > button:hover { opacity: .88; box-shadow: 0 5px 16px rgba(227,30,36,.4); }

    /* ── Tabs ────────────────────────────────── */
    [data-testid="stTabs"] [role="tab"] {
        font-weight: 600; font-size: .9rem;
        color: #555;
        border-bottom: 3px solid transparent;
        padding: 10px 20px;
    }
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        color: #e31e24 !important;
        border-bottom: 3px solid #e31e24 !important;
    }

    /* ── Info / metric ───────────────────────── */
    [data-testid="stMetric"] { background: #fff; border-radius: 10px; padding: 12px; box-shadow: 0 2px 8px rgba(0,0,0,.06); }

    /* ── Input widgets ───────────────────────── */
    .stSelectbox label, .stDateInput label, .stNumberInput label { font-weight: 600; color: #1a1a2e; }
    
    /* ── Success / warning / error ───────────── */
    .stAlert { border-radius: 8px; }

    /* ── Download button ─────────────────────── */
    .stDownloadButton > button {
        background: linear-gradient(90deg, #0f3460, #16213e);
        color: #fff !important;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        padding: 10px 22px;
    }
    
    /* sidebar nav buttons */
    .sidebar-nav-btn {
        display: block;
        width: 100%;
        padding: 12px 16px;
        margin: 6px 0;
        background: rgba(255,255,255,0.08);
        border-radius: 8px;
        color: #fff;
        font-weight: 600;
        text-align: left;
        border: none;
        cursor: pointer;
        transition: background .2s;
    }
    .sidebar-nav-btn:hover { background: rgba(227,30,36,0.35); }
    .sidebar-nav-btn.active { background: rgba(227,30,36,0.6); }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# LOGO HELPER
# ─────────────────────────────────────────────
def get_logo_b64(path="LOGO1.png"):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return None

logo_b64 = get_logo_b64()

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:52px;border-radius:6px;"/>' if logo_b64 else ""
st.markdown(f"""
<div class="cfg-header">
    {logo_html}
    <div>
        <h1>Système de Valorisation OPCVM</h1>
        <p>Conforme à la Circulaire CDVM N°02/04 – Valorisation des titres de créances à taux fixe remboursables in fine</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
if "bam_curve" not in st.session_state:
    st.session_state.bam_curve = None
if "df_titres" not in st.session_state:
    st.session_state.df_titres = None
if "df_results" not in st.session_state:
    st.session_state.df_results = None
if "bam_date" not in st.session_state:
    st.session_state.bam_date = None

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    if logo_b64:
        st.markdown(f'<div style="text-align:center;padding:10px 0 20px"><img src="data:image/png;base64,{logo_b64}" style="height:44px;"/></div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 🗂️ Navigation")
    page = st.radio(
        "",
        ["🏠 Accueil & Courbe BAM", "📊 Valorisation", "📈 Analyse & Graphiques", "📋 Résultats & Export"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.markdown("### ℹ️ Informations")
    st.markdown(f"""
    <div style="font-size:.82rem; color:#ccc; line-height:1.6">
    <b>Date d'évaluation :</b><br/>
    {datetime.date.today().strftime('%d/%m/%Y')}<br/><br/>
    <b>Référentiel :</b><br/>
    Circulaire CDVM N°02/04<br/><br/>
    <b>Source taux :</b><br/>
    Bank Al-Maghrib (BAM)
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("""
    <div style="font-size:.75rem;color:#888;text-align:center">
    CFG Bank – Direction Gestion d'Actifs<br/>
    © 2024 – Tous droits réservés
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

# BAM standard maturities in years
BAM_MATURITIES_YEARS = [13/52, 26/52, 52/52, 2, 5, 10, 15, 20, 25, 30]
BAM_MATURITIES_WEEKS = ["13 semaines", "26 semaines", "52 semaines", "2 ans", "5 ans", "10 ans", "15 ans", "20 ans", "25 ans", "30 ans"]

def fetch_bam_curve():
    """Fetch BAM yield curve from Bank Al-Maghrib website."""
    url = "https://www.bkam.ma/Marches/Principaux-indicateurs/Marche-obligataire/Marche-des-bons-du-tresor/Marche-secondaire/Taux-de-reference-des-bons-du-tresor"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "fr-FR,fr;q=0.9"
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            return None, "Table non trouvée sur le site BAM"
        rows = table.find_all("tr")
        data = []
        for row in rows:
            cols = [td.get_text(strip=True).replace(",", ".") for td in row.find_all(["td","th"])]
            if cols:
                data.append(cols)
        if len(data) < 2:
            return None, "Données insuffisantes dans la table BAM"
        df = pd.DataFrame(data[1:], columns=data[0])
        return df, None
    except Exception as e:
        return None, str(e)

def parse_bam_curve_manual(text_input):
    """Parse manually entered BAM curve."""
    lines = [l.strip() for l in text_input.strip().split('\n') if l.strip()]
    mats = []
    rates = []
    for line in lines:
        parts = line.replace(',', '.').split()
        if len(parts) >= 2:
            try:
                r = float(parts[-1]) / 100 if float(parts[-1]) > 1 else float(parts[-1])
                mats.append(' '.join(parts[:-1]))
                rates.append(r)
            except:
                pass
    return mats, rates

def interpolate_rate(maturity_days, curve_maturities_days, curve_rates):
    """
    Linear interpolation per Circulaire 02/04 Article 6.
    For maturities below the first point (8-13 weeks), use the first point.
    """
    curve_maturities_days = np.array(curve_maturities_days)
    curve_rates = np.array(curve_rates)
    
    idx = np.argsort(curve_maturities_days)
    curve_maturities_days = curve_maturities_days[idx]
    curve_rates = curve_rates[idx]
    
    if maturity_days <= curve_maturities_days[0]:
        return float(curve_rates[0])
    if maturity_days >= curve_maturities_days[-1]:
        return float(curve_rates[-1])
    
    for i in range(len(curve_maturities_days) - 1):
        if curve_maturities_days[i] <= maturity_days <= curve_maturities_days[i+1]:
            x0, x1 = curve_maturities_days[i], curve_maturities_days[i+1]
            y0, y1 = curve_rates[i], curve_rates[i+1]
            return float(y0 + (y1 - y0) * (maturity_days - x0) / (x1 - x0))
    return float(curve_rates[-1])

def valoriser_titre(row, eval_date, curve_maturities_days, curve_rates):
    """
    Valorise un titre selon la Circulaire CDVM 02/04.
    Returns (price, prix_pied_coupon, coupon_couru, taux_actualisation, methode, detail)
    """
    nature = str(row.get("Nature", "")).strip()
    nom    = str(row.get("TITRE", ""))
    
    date_emission = row.get("Date Emission")
    date_echeance = row.get("Date Échéance")
    tf   = float(row.get("Taux facial", 0) or 0)
    spread = float(row.get("SPREAD", 0) or 0)
    nominal = row.get("Nominal", 100000)
    
    # Parse nominal (can be string with spaces)
    try:
        if isinstance(nominal, str):
            nominal = float(nominal.replace(" ", "").replace(",", ".").replace("\xa0", ""))
        else:
            nominal = float(nominal)
    except:
        nominal = 100000

    nb_jours   = row.get("Nb jours", 0) or 0   # nj: days to next coupon
    nb_coupons = row.get("nombre de coupons", 1) or 1
    A          = row.get("A", 365) or 365
    prochain_coupon = row.get("Prochain coupon")

    # Compute Mi and Mr
    if isinstance(date_emission, datetime.datetime):
        date_emission = date_emission.date()
    if isinstance(date_echeance, datetime.datetime):
        date_echeance = date_echeance.date()
    if isinstance(eval_date, datetime.datetime):
        eval_date = eval_date.date()
    if isinstance(prochain_coupon, datetime.datetime):
        prochain_coupon = prochain_coupon.date()

    if date_emission is None or date_echeance is None:
        return None, None, None, None, "Données manquantes", {}

    Mi = (date_echeance - date_emission).days
    Mr = (date_echeance - eval_date).days

    if Mr <= 0:
        return None, None, None, None, "Échu", {}

    # Taux BAM interpolé pour Mr
    tr_bam = interpolate_rate(Mr, curve_maturities_days, curve_rates)
    tr = tr_bam + spread  # add risk/liquidity premium

    detail = {
        "Mi (jours)": Mi,
        "Mr (jours)": Mr,
        "Taux BAM interpolé": f"{tr_bam*100:.4f}%",
        "Spread": f"{spread*100:.4f}%",
        "Taux Actuali. (tr)": f"{tr*100:.4f}%",
        "Taux Facial (tf)": f"{tf*100:.4f}%",
        "Nominal": f"{nominal:,.2f}",
        "A": int(A),
    }

    price = None
    methode = ""
    coupon_couru = 0.0

    # ── FORMULE 1: Mi <= 365 ───────────────────────────────────────────
    if Mi <= 365:
        methode = "Formule (1) – Mi ≤ 365j"
        try:
            numerator = 1 + tf * Mi / 360
            denominator = 1 + tr * Mr / 360
            price = nominal * numerator / denominator
            detail["Formule"] = f"N × (1 + tf×Mi/360) / (1 + tr×Mr/360)"
        except Exception as e:
            return None, None, None, tr, f"Erreur: {e}", detail

    # ── FORMULE 2: Mi > 365 And Mr <= 365 ────────────────────────────
    elif Mi > 365 and Mr <= 365:
        methode = "Formule (2) – Mi > 365j & Mr ≤ 365j"
        try:
            # Check if it's a "ligne postérieure à un seul flux" case
            if nb_coupons <= 1 and ('Atypique' in nature or 'postérieure' in nature.lower()):
                # Formula (3)
                methode = "Formule (3) – Ligne postérieure 1 flux"
                numerator = 1 + tf * Mi / A
                denominator = 1 + tr * Mr / 360
                price = nominal * numerator / denominator
                detail["Formule"] = f"N × (1 + tf×Mi/A) / (1 + tr×Mr/360)"
            else:
                # Standard formula (2)
                numerator = 1 + tf
                denominator = 1 + tr * Mr / 360
                price = nominal * numerator / denominator
                detail["Formule"] = f"N × (1 + tf) / (1 + tr×Mr/360)"
            
            # Coupon couru
            if prochain_coupon:
                days_to_next = (prochain_coupon - eval_date).days
                days_in_period = A
                coupon_couru = nominal * tf * (1 - days_to_next / days_in_period)
        except Exception as e:
            return None, None, None, tr, f"Erreur: {e}", detail

    # ── FORMULE 4: Mi > 365 And Mr > 365 ─────────────────────────────
    else:
        methode = "Formule (4) – Mi > 365j & Mr > 365j"
        try:
            nj = int(nb_jours) if nb_jours else 0
            n  = int(nb_coupons) if nb_coupons else 1
            A  = int(A) if A else 365

            # Rebuild coupon schedule
            # First cash flow at nj days, then annually
            # Using formula (4.1) for normal line
            cash_flows = []
            # n coupons: at nj, nj+A, nj+2A, ..., nj+(n-1)A
            for i in range(1, n + 1):
                if i == n:
                    Fi = nominal * (1 + tf)  # last coupon + repayment
                else:
                    Fi = nominal * tf
                cash_flows.append(Fi)

            # Discount using formula (4)
            price = 0.0
            for i, Fi in enumerate(cash_flows):
                exponent = nj / A + i  # (nj/A) + (i-1) where i starts at 1 → use index
                if i == 0:
                    exponent = nj / A
                else:
                    exponent = nj / A + i
                price += Fi / ((1 + tr) ** exponent)
            
            detail["Formule"] = f"Somme[Fi / (1+tr)^(nj/A + i)] — {n} flux"
            detail["nj (j. prochain coupon)"] = nj
            detail["n (nb coupons)"] = n

            # Coupon couru
            if prochain_coupon and nj > 0:
                coupon_couru = nominal * tf * (1 - nj / A)
            
        except Exception as e:
            return None, None, None, tr, f"Erreur: {e}", detail

    if price is not None:
        prix_pied_coupon = price - coupon_couru
        return price, prix_pied_coupon, coupon_couru, tr, methode, detail
    return None, None, None, tr, methode, detail


def load_excel(file):
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    return pd.DataFrame(data)

# ─────────────────────────────────────────────
# PAGE 1 – ACCUEIL & COURBE BAM
# ─────────────────────────────────────────────
if "Accueil" in page:
    st.markdown('<div class="section-title">📡 Importation de la Courbe des Taux BAM</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.info("La courbe des taux est publiée quotidiennement par Bank Al-Maghrib. Vous pouvez l'importer automatiquement ou la saisir manuellement.")
        
        tab_auto, tab_manual, tab_csv = st.tabs(["🌐 Import Automatique BAM", "✏️ Saisie Manuelle", "📁 Import CSV/Excel"])
        
        with tab_auto:
            st.markdown("**Import direct depuis le site Bank Al-Maghrib**")
            st.warning("⚠️ Assurez-vous d'avoir accès à internet et au site www.bkam.ma")
            
            if st.button("🔄 Importer la courbe BAM maintenant", key="btn_auto"):
                with st.spinner("Connexion à Bank Al-Maghrib en cours..."):
                    df_bam, err = fetch_bam_curve()
                    if err:
                        st.error(f"Erreur lors de l'import : {err}")
                        st.info("💡 Utilisez la saisie manuelle ou l'import CSV en alternative.")
                    else:
                        st.success("✅ Courbe BAM importée avec succès !")
                        st.dataframe(df_bam)
                        # Try to parse
                        st.session_state.bam_date = datetime.date.today()
        
        with tab_manual:
            st.markdown("**Saisie manuelle de la courbe des taux BAM**")
            st.caption("Entrez chaque ligne sous la forme : `Maturité  Taux(%)` séparés par un espace ou tabulation")
            
            default_text = """13 semaines  2.55
26 semaines  2.60
52 semaines  2.65
2 ans        2.78
5 ans        3.10
10 ans       3.45
15 ans       3.80
20 ans       4.05
25 ans       4.20
30 ans       4.35"""
            
            manual_input = st.text_area(
                "Courbe des taux (Maturité  Taux%)",
                value=default_text,
                height=250,
                help="Taux en pourcentage. Ex: '5 ans 3.10' signifie 3.10%"
            )
            
            eval_date_manual = st.date_input(
                "Date d'évaluation (date de la courbe BAM)",
                value=datetime.date.today(),
                format="DD/MM/YYYY"
            )
            
            if st.button("✅ Valider la courbe saisie", key="btn_manual"):
                lines = [l.strip() for l in manual_input.strip().split('\n') if l.strip()]
                maturities_days = []
                rates = []
                mat_labels = []
                errors = []
                
                mat_map = {
                    "13 semaines": 13*7, "26 semaines": 26*7, "52 semaines": 52*7,
                    "1 an": 365, "2 ans": 2*365, "3 ans": 3*365, "4 ans": 4*365,
                    "5 ans": 5*365, "7 ans": 7*365, "10 ans": 10*365,
                    "15 ans": 15*365, "20 ans": 20*365, "25 ans": 25*365, "30 ans": 30*365
                }
                
                for line in lines:
                    parts = line.replace(',', '.').split()
                    if len(parts) < 2:
                        continue
                    try:
                        rate_val = float(parts[-1])
                        rate = rate_val / 100 if rate_val > 1 else rate_val
                        mat_str = ' '.join(parts[:-1]).lower()
                        
                        # Try to find days
                        days = None
                        for key, d in mat_map.items():
                            if key.lower() in mat_str or mat_str in key.lower():
                                days = d
                                mat_labels.append(key)
                                break
                        if days is None:
                            # Try to parse "X ans" or "X semaines"
                            if 'an' in mat_str:
                                n = float(mat_str.replace('ans','').replace('an','').strip())
                                days = int(n * 365)
                                mat_labels.append(f"{n:.0f} ans")
                            elif 'sem' in mat_str:
                                n = float(mat_str.replace('semaines','').replace('semaine','').strip())
                                days = int(n * 7)
                                mat_labels.append(f"{n:.0f} semaines")
                        if days:
                            maturities_days.append(days)
                            rates.append(rate)
                    except Exception as e:
                        errors.append(f"Ligne ignorée: {line} ({e})")
                
                if errors:
                    for e in errors:
                        st.warning(e)
                
                if len(rates) >= 2:
                    st.session_state.bam_curve = {
                        "maturities_days": maturities_days,
                        "rates": rates,
                        "labels": mat_labels,
                        "date": eval_date_manual
                    }
                    st.session_state.bam_date = eval_date_manual
                    st.success(f"✅ Courbe validée : {len(rates)} points de maturité enregistrés")
                else:
                    st.error("❌ Impossible de parser la courbe. Vérifiez le format.")
        
        with tab_csv:
            st.markdown("**Import depuis fichier CSV ou Excel (format BAM)**")
            st.caption("Le fichier doit contenir une colonne 'Maturité' et une colonne 'Taux' ou 'Rendement'")
            
            uploaded_bam = st.file_uploader("Choisir le fichier BAM (CSV ou XLSX)", type=["csv", "xlsx", "xls"])
            eval_date_csv = st.date_input("Date d'évaluation", value=datetime.date.today(), format="DD/MM/YYYY", key="csv_date")
            
            if uploaded_bam and st.button("📥 Charger la courbe", key="btn_csv"):
                try:
                    if uploaded_bam.name.endswith(".csv"):
                        df_csv = pd.read_csv(uploaded_bam, sep=None, engine='python', decimal=',')
                    else:
                        df_csv = pd.read_excel(uploaded_bam)
                    
                    st.dataframe(df_csv.head(15))
                    
                    # Try auto-detect columns
                    cols = [c.lower() for c in df_csv.columns]
                    mat_col = next((df_csv.columns[i] for i, c in enumerate(cols) if 'mat' in c or 'dur' in c or 'ech' in c), None)
                    rate_col = next((df_csv.columns[i] for i, c in enumerate(cols) if 'taux' in c or 'rend' in c or 'rate' in c), None)
                    
                    if mat_col and rate_col:
                        st.success(f"Colonnes détectées: Maturité='{mat_col}', Taux='{rate_col}'")
                        # Further processing would go here
                        st.info("Fonctionnalité de parsing automatique CSV en cours de développement. Utilisez la saisie manuelle.")
                    else:
                        st.warning("Colonnes non détectées automatiquement.")
                        col_options = list(df_csv.columns)
                        mat_col = st.selectbox("Colonne Maturité", col_options)
                        rate_col = st.selectbox("Colonne Taux", col_options)
                except Exception as e:
                    st.error(f"Erreur : {e}")

    with col2:
        st.markdown('<div class="section-title">📊 Courbe Actuelle</div>', unsafe_allow_html=True)
        
        if st.session_state.bam_curve:
            curve = st.session_state.bam_curve
            st.success(f"✅ Courbe chargée\n\nDate: **{curve['date'].strftime('%d/%m/%Y')}**\n\nPoints: **{len(curve['rates'])}**")
            
            df_curve = pd.DataFrame({
                "Maturité": curve.get("labels", [f"{d//365:.1f}a" for d in curve["maturities_days"]]),
                "Jours": curve["maturities_days"],
                "Taux (%)": [f"{r*100:.4f}%" for r in curve["rates"]]
            })
            st.dataframe(df_curve, use_container_width=True, hide_index=True)
            
            # Mini chart
            fig = go.Figure()
            x_labels = curve.get("labels", [str(d) for d in curve["maturities_days"]])
            fig.add_trace(go.Scatter(
                x=x_labels,
                y=[r*100 for r in curve["rates"]],
                mode='lines+markers',
                line=dict(color='#e31e24', width=2.5),
                marker=dict(size=7, color='#1a1a2e'),
                name="Taux BAM"
            ))
            fig.update_layout(
                title=f"Courbe BAM – {curve['date'].strftime('%d/%m/%Y')}",
                xaxis_title="Maturité",
                yaxis_title="Taux (%)",
                height=280,
                margin=dict(l=30, r=20, t=40, b=40),
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(size=10)
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("⚠️ Aucune courbe chargée.\nVeuillez importer ou saisir la courbe BAM.")

    # File upload section
    st.markdown('<div class="section-title">📁 Chargement du Fichier de Données</div>', unsafe_allow_html=True)
    
    col_up1, col_up2 = st.columns([3, 1])
    with col_up1:
        uploaded_file = st.file_uploader(
            "Glissez-déposez votre fichier Excel de valorisation (DATA_VALORISATION.xlsx)",
            type=["xlsx", "xls"],
            help="Le fichier doit contenir les colonnes: Nature, TITRE, CODE ISIN, Date Emission, Date Échéance, Taux facial, SPREAD, Nominal, etc."
        )
        if uploaded_file:
            with st.spinner("Chargement des données..."):
                df = load_excel(uploaded_file)
                if df is not None:
                    st.session_state.df_titres = df
                    st.success(f"✅ Fichier chargé : **{len(df)} titres** importés")
                    st.dataframe(df[["TITRE", "CODE ISIN", "Nature", "Date Emission", "Date Échéance", "Taux facial", "SPREAD", "Nominal"]].head(10), 
                                use_container_width=True, hide_index=True)
    
    with col_up2:
        if st.session_state.df_titres is not None:
            df = st.session_state.df_titres
            natures = df["Nature"].value_counts()
            st.metric("Titres chargés", len(df))
            
            fig_donut = go.Figure(data=[go.Pie(
                labels=[n[:20]+"..." if len(str(n))>20 else str(n) for n in natures.index],
                values=natures.values,
                hole=0.55,
                marker_colors=['#e31e24','#1a1a2e','#0f3460','#c0141a','#16213e','#aaa']
            )])
            fig_donut.update_layout(
                height=220, margin=dict(l=0, r=0, t=10, b=0),
                showlegend=False, paper_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig_donut, use_container_width=True)

# ─────────────────────────────────────────────
# PAGE 2 – VALORISATION
# ─────────────────────────────────────────────
elif "Valorisation" in page:
    st.markdown('<div class="section-title">⚙️ Paramètres de Valorisation</div>', unsafe_allow_html=True)
    
    if st.session_state.df_titres is None:
        st.warning("⚠️ Veuillez d'abord charger votre fichier de données dans l'onglet 'Accueil'.")
        st.stop()
    if st.session_state.bam_curve is None:
        st.warning("⚠️ Veuillez d'abord charger la courbe des taux BAM dans l'onglet 'Accueil'.")
        st.stop()

    col1, col2, col3 = st.columns(3)
    with col1:
        eval_date = st.date_input(
            "📅 Date d'évaluation",
            value=datetime.date.today(),
            format="DD/MM/YYYY"
        )
    with col2:
        filter_nature = st.multiselect(
            "🔍 Filtrer par Nature",
            options=["Tous"] + list(st.session_state.df_titres["Nature"].dropna().unique()),
            default=["Tous"]
        )
    with col3:
        filter_modele = st.multiselect(
            "🔍 Filtrer par Modèle",
            options=["Tous"] + list(st.session_state.df_titres["MODELE"].dropna().unique()),
            default=["Tous"]
        )
    
    st.markdown("---")
    
    df = st.session_state.df_titres.copy()
    if "Tous" not in filter_nature and filter_nature:
        df = df[df["Nature"].isin(filter_nature)]
    if "Tous" not in filter_modele and filter_modele:
        df = df[df["MODELE"].isin(filter_modele)]
    
    st.info(f"🔢 **{len(df)} titres** sélectionnés pour la valorisation")
    
    col_btn1, col_btn2 = st.columns([1, 4])
    with col_btn1:
        run = st.button("🚀 Lancer la Valorisation", key="btn_run")
    
    if run:
        curve = st.session_state.bam_curve
        mats = curve["maturities_days"]
        rates = curve["rates"]
        
        results = []
        progress = st.progress(0, text="Valorisation en cours...")
        total = len(df)
        
        for i, (_, row) in enumerate(df.iterrows()):
            price, ppc, cc, tr, methode, detail = valoriser_titre(
                row.to_dict(), eval_date, mats, rates
            )
            
            nominal = row.get("Nominal", 100000)
            try:
                if isinstance(nominal, str):
                    nominal = float(nominal.replace(" ", "").replace(",", ".").replace("\xa0", ""))
                else:
                    nominal = float(nominal) if nominal else 100000
            except:
                nominal = 100000
            
            # Compute Mr for display
            de = row.get("Date Emission")
            dec = row.get("Date Échéance")
            if isinstance(de, datetime.datetime): de = de.date()
            if isinstance(dec, datetime.datetime): dec = dec.date()
            Mr = (dec - eval_date).days if dec and eval_date else None
            Mi = (dec - de).days if dec and de else None
            
            results.append({
                "TITRE": row.get("TITRE", ""),
                "CODE ISIN": row.get("CODE ISIN", ""),
                "Nature": row.get("Nature", ""),
                "MODELE": row.get("MODELE", ""),
                "Date Emission": de,
                "Date Échéance": dec,
                "Mi (j)": Mi,
                "Mr (j)": Mr,
                "Taux Facial (%)": float(row.get("Taux facial", 0) or 0) * 100,
                "Spread (%)": float(row.get("SPREAD", 0) or 0) * 100,
                "Taux Act. (%)": float(tr)*100 if tr else None,
                "Nominal": nominal,
                "Prix (%) ": round(price / nominal * 100, 6) if price and nominal else None,
                "Prix Absolu": round(price, 4) if price else None,
                "Coupon Couru": round(cc, 4) if cc else 0,
                "Prix Pied Coupon": round(ppc, 4) if ppc else None,
                "Méthode": methode,
                "Statut": "✅ OK" if price is not None else "❌ Erreur"
            })
            progress.progress((i + 1) / total, text=f"Valorisation : {i+1}/{total} titres traités")
        
        progress.empty()
        df_res = pd.DataFrame(results)
        st.session_state.df_results = df_res
        st.success(f"✅ Valorisation terminée ! **{len(df_res[df_res['Statut']=='✅ OK'])}** titres valorisés avec succès.")
        
        # Summary KPIs
        ok = df_res[df_res["Statut"] == "✅ OK"]
        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi-card">
                <div class="kpi-label">Titres Valorisés</div>
                <div class="kpi-value">{len(ok)}</div>
                <div class="kpi-sub">sur {len(df_res)} total</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Valeur Totale</div>
                <div class="kpi-value">{ok['Prix Absolu'].sum():,.0f}</div>
                <div class="kpi-sub">MAD</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Prix Moyen (%)</div>
                <div class="kpi-value">{ok['Prix (%) '].mean():.4f}%</div>
                <div class="kpi-sub">du nominal</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Taux Act. Moyen</div>
                <div class="kpi-value">{ok['Taux Act. (%)'].mean():.4f}%</div>
                <div class="kpi-sub">pondéré</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.dataframe(df_res, use_container_width=True, hide_index=True)
    
    elif st.session_state.df_results is not None:
        df_res = st.session_state.df_results
        ok = df_res[df_res["Statut"] == "✅ OK"]
        st.markdown(f"""
        <div class="kpi-row">
            <div class="kpi-card">
                <div class="kpi-label">Titres Valorisés</div>
                <div class="kpi-value">{len(ok)}</div>
                <div class="kpi-sub">sur {len(df_res)} total</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Valeur Totale</div>
                <div class="kpi-value">{ok['Prix Absolu'].sum():,.0f}</div>
                <div class="kpi-sub">MAD</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Prix Moyen (%)</div>
                <div class="kpi-value">{ok['Prix (%) '].mean():.4f}%</div>
                <div class="kpi-sub">du nominal</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(df_res, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# PAGE 3 – ANALYSE & GRAPHIQUES
# ─────────────────────────────────────────────
elif "Analyse" in page:
    st.markdown('<div class="section-title">📈 Analyse des Résultats de Valorisation</div>', unsafe_allow_html=True)
    
    if st.session_state.df_results is None:
        st.warning("⚠️ Veuillez d'abord lancer la valorisation.")
        st.stop()
    
    df_res = st.session_state.df_results.copy()
    ok = df_res[df_res["Statut"] == "✅ OK"].copy()
    
    if ok.empty:
        st.error("Aucun titre valorisé avec succès.")
        st.stop()

    # Courbe BAM
    if st.session_state.bam_curve:
        curve = st.session_state.bam_curve
        st.markdown('<div class="section-title">📉 Courbe des Taux BAM</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            fig_bam = go.Figure()
            x_labels = curve.get("labels", [str(d) for d in curve["maturities_days"]])
            fig_bam.add_trace(go.Scatter(
                x=x_labels,
                y=[r*100 for r in curve["rates"]],
                mode='lines+markers',
                line=dict(color='#e31e24', width=3),
                marker=dict(size=9, color='#1a1a2e', line=dict(color='#e31e24', width=2)),
                fill='tozeroy',
                fillcolor='rgba(227,30,36,0.08)',
                name="Taux BAM"
            ))
            fig_bam.update_layout(
                title=f"Courbe des Taux de Référence BAM – {curve['date'].strftime('%d/%m/%Y')}",
                xaxis_title="Maturité",
                yaxis_title="Taux (%)",
                height=380,
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(family="Arial", size=11),
                legend=dict(orientation="h")
            )
            st.plotly_chart(fig_bam, use_container_width=True)
        
        with col2:
            # Taux actualisations vs maturités résiduelles
            fig_scatter = go.Figure()
            fig_scatter.add_trace(go.Scatter(
                x=ok["Mr (j)"],
                y=ok["Taux Act. (%)"],
                mode='markers',
                marker=dict(
                    size=9,
                    color=ok["Prix (%) "],
                    colorscale='RdBu',
                    showscale=True,
                    colorbar=dict(title="Prix %"),
                    line=dict(color='#1a1a2e', width=0.5)
                ),
                text=ok["TITRE"],
                hovertemplate="<b>%{text}</b><br>Mr: %{x}j<br>Taux: %{y:.4f}%<extra></extra>"
            ))
            fig_scatter.update_layout(
                title="Taux d'actualisation vs Maturité Résiduelle",
                xaxis_title="Maturité Résiduelle (jours)",
                yaxis_title="Taux d'Actualisation (%)",
                height=380,
                plot_bgcolor='white',
                paper_bgcolor='white'
            )
            st.plotly_chart(fig_scatter, use_container_width=True)
    
    # Distribution des prix
    col1, col2 = st.columns(2)
    
    with col1:
        fig_hist = go.Figure()
        fig_hist.add_trace(go.Histogram(
            x=ok["Prix (%) "],
            nbinsx=30,
            marker_color='#e31e24',
            marker_line_color='white',
            marker_line_width=0.5,
            opacity=0.85,
            name="Distribution Prix"
        ))
        fig_hist.update_layout(
            title="Distribution des Prix (% du Nominal)",
            xaxis_title="Prix (%)",
            yaxis_title="Nombre de titres",
            height=360,
            plot_bgcolor='white',
            paper_bgcolor='white'
        )
        st.plotly_chart(fig_hist, use_container_width=True)
    
    with col2:
        # Répartition par méthode
        methode_counts = ok["Méthode"].value_counts()
        fig_pie = go.Figure(data=[go.Pie(
            labels=[m[:35]+"..." if len(m) > 35 else m for m in methode_counts.index],
            values=methode_counts.values,
            hole=0.4,
            marker_colors=['#e31e24', '#1a1a2e', '#0f3460', '#c0141a', '#aaa']
        )])
        fig_pie.update_layout(
            title="Répartition par Formule de Valorisation",
            height=360,
            paper_bgcolor='white'
        )
        st.plotly_chart(fig_pie, use_container_width=True)
    
    # Prix par titre (top/bottom)
    st.markdown('<div class="section-title">🏆 Classement des Titres par Prix</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        top10 = ok.nlargest(10, "Prix (%) ")[["TITRE", "Prix (%) ", "Taux Act. (%)", "Mr (j)"]]
        fig_top = go.Figure(go.Bar(
            x=top10["Prix (%) "],
            y=[t[:25]+"..." if len(t)>25 else t for t in top10["TITRE"]],
            orientation='h',
            marker_color='#e31e24',
            text=[f"{v:.2f}%" for v in top10["Prix (%) "]],
            textposition='inside'
        ))
        fig_top.update_layout(
            title="Top 10 – Prix les Plus Élevés",
            xaxis_title="Prix (%)",
            height=380,
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(l=180)
        )
        st.plotly_chart(fig_top, use_container_width=True)
    
    with col2:
        bot10 = ok.nsmallest(10, "Prix (%) ")[["TITRE", "Prix (%) ", "Taux Act. (%)", "Mr (j)"]]
        fig_bot = go.Figure(go.Bar(
            x=bot10["Prix (%) "],
            y=[t[:25]+"..." if len(t)>25 else t for t in bot10["TITRE"]],
            orientation='h',
            marker_color='#1a1a2e',
            text=[f"{v:.2f}%" for v in bot10["Prix (%) "]],
            textposition='inside'
        ))
        fig_bot.update_layout(
            title="Top 10 – Prix les Plus Faibles",
            xaxis_title="Prix (%)",
            height=380,
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(l=180)
        )
        st.plotly_chart(fig_bot, use_container_width=True)
    
    # Spread analysis
    st.markdown('<div class="section-title">📊 Analyse des Spreads</div>', unsafe_allow_html=True)
    fig_spread = px.scatter(
        ok,
        x="Spread (%)",
        y="Prix (%) ",
        color="Méthode",
        size="Nominal",
        hover_name="TITRE",
        color_discrete_sequence=['#e31e24','#1a1a2e','#0f3460','#c0141a','#aaa'],
        title="Spread vs Prix par Titre"
    )
    fig_spread.update_layout(height=400, plot_bgcolor='white', paper_bgcolor='white')
    st.plotly_chart(fig_spread, use_container_width=True)

# ─────────────────────────────────────────────
# PAGE 4 – RÉSULTATS & EXPORT
# ─────────────────────────────────────────────
elif "Résultats" in page:
    st.markdown('<div class="section-title">📋 Résultats Détaillés & Export</div>', unsafe_allow_html=True)
    
    if st.session_state.df_results is None:
        st.warning("⚠️ Veuillez d'abord lancer la valorisation.")
        st.stop()
    
    df_res = st.session_state.df_results.copy()
    
    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        search = st.text_input("🔍 Recherche (titre ou ISIN)", "")
    with col2:
        statut_filter = st.selectbox("Statut", ["Tous", "✅ OK uniquement", "❌ Erreurs uniquement"])
    with col3:
        sort_col = st.selectbox("Trier par", ["TITRE", "Prix (%) ", "Taux Act. (%)", "Mr (j)", "Spread (%)"])
    
    df_display = df_res.copy()
    if search:
        df_display = df_display[
            df_display["TITRE"].str.contains(search, case=False, na=False) |
            df_display["CODE ISIN"].str.contains(search, case=False, na=False)
        ]
    if statut_filter == "✅ OK uniquement":
        df_display = df_display[df_display["Statut"] == "✅ OK"]
    elif statut_filter == "❌ Erreurs uniquement":
        df_display = df_display[df_display["Statut"] != "✅ OK"]
    
    df_display = df_display.sort_values(sort_col, ascending=False, na_position='last')
    
    ok = df_display[df_display["Statut"] == "✅ OK"]
    
    # KPIs
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total titres", len(df_display))
    col2.metric("Valorisés ✅", len(ok))
    col3.metric("Erreurs ❌", len(df_display) - len(ok))
    if len(ok) > 0:
        col4.metric("Valeur Totale", f"{ok['Prix Absolu'].sum():,.0f} MAD")
        col5.metric("Prix Moyen", f"{ok['Prix (%) '].mean():.4f}%")
    
    st.markdown("---")
    
    # Color-coded table
    st.dataframe(
        df_display.style.apply(
            lambda row: ['background-color: #fff5f5' if row['Statut'] != '✅ OK' else '' for _ in row],
            axis=1
        ).format({
            "Taux Facial (%)": "{:.4f}%",
            "Spread (%)": "{:.4f}%",
            "Taux Act. (%)": "{:.4f}%",
            "Prix (%) ": "{:.6f}%",
            "Prix Absolu": "{:,.4f}",
            "Coupon Couru": "{:,.4f}",
            "Prix Pied Coupon": "{:,.4f}",
            "Nominal": "{:,.2f}",
        }, na_rep="-"),
        use_container_width=True,
        height=500
    )
    
    # Detail panel
    st.markdown('<div class="section-title">🔎 Détail par Titre</div>', unsafe_allow_html=True)
    titre_sel = st.selectbox("Sélectionner un titre", df_display["TITRE"].tolist())
    if titre_sel:
        row_sel = df_display[df_display["TITRE"] == titre_sel].iloc[0]
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Prix (%)", f"{row_sel['Prix (%) ']:.6f}%" if row_sel['Prix (%) '] else "N/A")
        col2.metric("Prix Absolu", f"{row_sel['Prix Absolu']:,.4f} MAD" if row_sel['Prix Absolu'] else "N/A")
        col3.metric("Coupon Couru", f"{row_sel['Coupon Couru']:,.4f}" if row_sel['Coupon Couru'] else "0")
        col4.metric("Taux Actualisation", f"{row_sel['Taux Act. (%)']:.4f}%" if row_sel['Taux Act. (%)'] else "N/A")
        
        col5, col6, col7, col8 = st.columns(4)
        col5.metric("Méthode", row_sel['Méthode'])
        col6.metric("Mr (jours)", f"{row_sel['Mr (j)']}" if row_sel['Mr (j)'] else "N/A")
        col7.metric("Taux Facial", f"{row_sel['Taux Facial (%)']:.4f}%")
        col8.metric("Spread", f"{row_sel['Spread (%)']:.4f}%")
    
    st.markdown("---")
    
    # Export section
    st.markdown('<div class="section-title">💾 Export des Résultats</div>', unsafe_allow_html=True)
    
    col_e1, col_e2, col_e3 = st.columns(3)
    
    with col_e1:
        # Excel export
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_res.to_excel(writer, sheet_name="Valorisation", index=False)
            if st.session_state.bam_curve:
                curve = st.session_state.bam_curve
                df_bam_exp = pd.DataFrame({
                    "Maturité": curve.get("labels", [f"{d}j" for d in curve["maturities_days"]]),
                    "Jours": curve["maturities_days"],
                    "Taux (%)": [r*100 for r in curve["rates"]]
                })
                df_bam_exp.to_excel(writer, sheet_name="Courbe BAM", index=False)
            # Summary sheet
            if len(df_res[df_res["Statut"]=="✅ OK"]) > 0:
                ok_exp = df_res[df_res["Statut"]=="✅ OK"]
                summary = pd.DataFrame({
                    "Indicateur": ["Nb titres valorisés", "Valeur Totale (MAD)", "Prix Moyen (%)", "Taux Act. Moyen (%)", "Date valorisation"],
                    "Valeur": [
                        len(ok_exp),
                        f"{ok_exp['Prix Absolu'].sum():,.2f}",
                        f"{ok_exp['Prix (%) '].mean():.6f}%",
                        f"{ok_exp['Taux Act. (%)'].mean():.4f}%",
                        datetime.date.today().strftime('%d/%m/%Y')
                    ]
                })
                summary.to_excel(writer, sheet_name="Résumé", index=False)
        
        st.download_button(
            label="📥 Télécharger Excel Complet",
            data=buf.getvalue(),
            file_name=f"Valorisation_OPCVM_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col_e2:
        # CSV export
        csv_buf = io.StringIO()
        df_res.to_csv(csv_buf, index=False, sep=";", decimal=",", encoding='utf-8-sig')
        st.download_button(
            label="📥 Télécharger CSV",
            data=csv_buf.getvalue().encode('utf-8-sig'),
            file_name=f"Valorisation_OPCVM_{datetime.date.today().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
    
    with col_e3:
        # OK only
        buf2 = io.BytesIO()
        ok_only = df_res[df_res["Statut"] == "✅ OK"]
        with pd.ExcelWriter(buf2, engine='openpyxl') as writer:
            ok_only.to_excel(writer, sheet_name="Valorisation OK", index=False)
        st.download_button(
            label="📥 Export Titres OK seulement",
            data=buf2.getvalue(),
            file_name=f"Valorisation_OK_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Errors details
    errors = df_res[df_res["Statut"] != "✅ OK"]
    if len(errors) > 0:
        with st.expander(f"⚠️ Détail des erreurs ({len(errors)} titres)", expanded=False):
            st.dataframe(errors[["TITRE", "CODE ISIN", "Nature", "Méthode", "Statut"]], 
                        use_container_width=True, hide_index=True)
